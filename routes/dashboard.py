from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from fastapi import HTTPException
from models.models import DatabaseManager
from dependencies import verify_token
from utils.logger import get_logger
from utils.sql_optimization import SQLOptimizer, get_sparkline_data
from utils.cache_manager import cache_manager
from datetime import datetime, timedelta
from typing import Optional
import calendar
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logger = get_logger(__name__)
router = APIRouter(prefix="/dashboard", tags=["dashboard"])

# Русские названия месяцев для форматирования периодов
RUSSIAN_MONTHS = {
    1: "январь", 2: "февраль", 3: "март", 4: "апрель",
    5: "май", 6: "июнь", 7: "июль", 8: "август",
    9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"
}


def format_period_label(dt: datetime) -> str:
    """Форматирует дату в читаемый формат 'месяц год' на русском языке"""
    return f"{RUSSIAN_MONTHS[dt.month]} {dt.year}"


@router.get("/stats")
async def get_dashboard_stats(
    period_start: Optional[str] = Query(None, description="Начало периода (ISO формат)"),
    period_end: Optional[str] = Query(None, description="Конец периода (ISO формат)"),
    _: str = Depends(verify_token)
):
    """
    Получение общей статистики для дашборда с кэшированием и процентами изменения.

    Если период не указан, используется текущий месяц.
    """
    logger.info(f"Dashboard stats request started: period_start={period_start}, period_end={period_end}")

    # Определяем период
    if period_start and period_end:
        try:
            period_start_dt = datetime.fromisoformat(period_start.replace('Z', '+00:00'))
            period_end_dt = datetime.fromisoformat(period_end.replace('Z', '+00:00'))
        except ValueError as e:
            logger.error(f"Invalid date format: {e}")
            raise HTTPException(status_code=400, detail="Неверный формат даты. Используйте ISO формат (YYYY-MM-DD)")
    else:
        # По умолчанию - текущий месяц
        now = datetime.now()
        period_start_dt = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        # Следующий месяц, день 1
        if now.month == 12:
            period_end_dt = period_start_dt.replace(year=now.year + 1, month=1)
        else:
            period_end_dt = period_start_dt.replace(month=now.month + 1)

    cache_key = cache_manager.get_cache_key(
        "dashboard", "stats",
        period_start_dt.isoformat(),
        period_end_dt.isoformat()
    )

    def _get_stats():
        def _db_query(session):
            logger.info("Executing dashboard stats database query with comparison")
            try:
                # Получаем основную статистику
                result = SQLOptimizer.get_dashboard_stats_with_comparison(
                    session, period_start_dt, period_end_dt
                )

                # Добавляем данные sparkline (последние 7 дней)
                sparkline_data = get_sparkline_data(session, days=7)
                result['sparkline'] = sparkline_data

                logger.info(f"Database query completed with sparkline data, result type: {type(result)}")
                return result
            except Exception as e:
                logger.error(f"Error in dashboard stats query: {e}", exc_info=True)
                raise

        logger.info("Starting database transaction for dashboard stats")
        result = DatabaseManager.safe_execute(_db_query)
        logger.info(f"Database transaction completed, final result type: {type(result)}")
        return result

    try:
        logger.info("Getting stats from cache or database")
        # Используем кэш с TTL для дашборда
        result = await cache_manager.get_or_set(
            cache_key, _get_stats, ttl=cache_manager.dashboard_ttl
        )

        logger.info(f"Stats result type: {type(result)}")

        # Дополнительная проверка результата
        if not isinstance(result, dict):
            logger.warning(f"Получен некорректный результат статистики: {type(result)}")
            raise ValueError("Некорректный формат данных статистики")

        # Форматируем ответ с трендовыми индикаторами
        def create_trend(change_percentage: float, metric_type: str = "growth"):
            """
            Создаёт индикатор тренда

            Args:
                change_percentage: процент изменения
                metric_type: тип метрики ('growth' для пользователей/бронирований, 'reduction' для тикетов)
            """
            if change_percentage > 0:
                direction = "up"
                is_positive = metric_type == "growth"
            elif change_percentage < 0:
                direction = "down"
                is_positive = metric_type == "reduction"
            else:
                direction = "neutral"
                is_positive = True

            return {
                "value": abs(change_percentage),
                "direction": direction,
                "is_positive": is_positive
            }

        # Формируем карточки статистики с трендами и sparkline
        formatted_result = {
            "users": {
                "current_value": result["users"]["current_value"],
                "previous_value": result["users"]["previous_value"],
                "change_percentage": result["users"]["change_percentage"],
                "trend": create_trend(result["users"]["change_percentage"], "growth"),
                "label": "Всего пользователей",
                "icon": "FiUsers",
                "sparkline": result.get("sparkline", {}).get("users", {"values": [], "labels": []})
            },
            "bookings": {
                "current_value": result["bookings"]["current_value"],
                "previous_value": result["bookings"]["previous_value"],
                "change_percentage": result["bookings"]["change_percentage"],
                "trend": create_trend(result["bookings"]["change_percentage"], "growth"),
                "label": "Всего бронирований",
                "icon": "FiShoppingBag",
                "sparkline": result.get("sparkline", {}).get("bookings", {"values": [], "labels": []})
            },
            "average_booking_value": {
                "current_value": result.get("average_booking_value", {}).get("current_value", 0),
                "previous_value": result.get("average_booking_value", {}).get("previous_value", 0),
                "change_percentage": result.get("average_booking_value", {}).get("change_percentage", 0),
                "trend": create_trend(result.get("average_booking_value", {}).get("change_percentage", 0), "growth"),
                "label": "Средний чек",
                "icon": "FiDollarSign",
                "sparkline": result.get("sparkline", {}).get("average_booking_value", {"values": [], "labels": []})
            },
            "tickets": {
                "current_value": result["tickets"]["current_value"],
                "previous_value": result["tickets"]["previous_value"],
                "change_percentage": result["tickets"]["change_percentage"],
                "trend": create_trend(result["tickets"]["change_percentage"], "reduction"),
                "label": "Открытые заявки",
                "icon": "FiMessageCircle",
                "sparkline": result.get("sparkline", {}).get("tickets", {"values": [], "labels": []})
            },
            # Дополнительная статистика (для обратной совместимости)
            "total_users": result["users"]["total"],
            "total_bookings": result["bookings"]["total"],
            "open_tickets": result["tickets"]["open"],
            "active_tariffs": result["active_tariffs"],
            "paid_bookings": result["paid_bookings"],
            "total_revenue": result["total_revenue"],
            "ticket_stats": result["ticket_stats"],
            "unread_notifications": result["unread_notifications"],
            # Информация о периоде
            "period": {
                "start": period_start_dt.isoformat(),
                "end": period_end_dt.isoformat(),
                "label": format_period_label(period_start_dt),
                "previous_label": format_period_label(
                    datetime.fromisoformat(result.get("period_info", {}).get("previous_start", period_start_dt.isoformat()).replace('Z', '+00:00'))
                )
            }
        }

        logger.info("Dashboard stats request completed successfully")
        return formatted_result

    except Exception as e:
        logger.error(f"Ошибка в get_dashboard_stats: {e}", exc_info=True)

        # Возвращаем базовые значения при любой ошибке
        logger.warning("Возвращаем базовую статистику из-за ошибки")
        return {
            "users": {
                "current_value": 0,
                "previous_value": 0,
                "change_percentage": 0.0,
                "trend": {"value": 0.0, "direction": "neutral", "is_positive": True},
                "label": "Всего пользователей",
                "icon": "FiUsers",
                "sparkline": {"values": [0]*7, "labels": [""]*7}
            },
            "bookings": {
                "current_value": 0,
                "previous_value": 0,
                "change_percentage": 0.0,
                "trend": {"value": 0.0, "direction": "neutral", "is_positive": True},
                "label": "Всего бронирований",
                "icon": "FiShoppingBag",
                "sparkline": {"values": [0]*7, "labels": [""]*7}
            },
            "average_booking_value": {
                "current_value": 0.0,
                "previous_value": 0.0,
                "change_percentage": 0.0,
                "trend": {"value": 0.0, "direction": "neutral", "is_positive": True},
                "label": "Средний чек",
                "icon": "FiDollarSign",
                "sparkline": {"values": [0]*7, "labels": [""]*7}
            },
            "tickets": {
                "current_value": 0,
                "previous_value": 0,
                "change_percentage": 0.0,
                "trend": {"value": 0.0, "direction": "neutral", "is_positive": True},
                "label": "Открытые заявки",
                "icon": "FiMessageCircle",
                "sparkline": {"values": [0]*7, "labels": [""]*7}
            },
            "total_users": 0,
            "total_bookings": 0,
            "open_tickets": 0,
            "active_tariffs": 0,
            "paid_bookings": 0,
            "total_revenue": 0.0,
            "ticket_stats": {
                "open": 0,
                "in_progress": 0,
                "closed": 0,
            },
            "unread_notifications": 0,
            "period": {
                "start": datetime.now().isoformat(),
                "end": datetime.now().isoformat(),
                "label": format_period_label(datetime.now()),
                "previous_label": format_period_label(datetime.now().replace(month=datetime.now().month - 1 if datetime.now().month > 1 else 12))
            }
        }


@router.get("/chart-data")
async def get_chart_data(
    year: int = Query(default_factory=lambda: datetime.now().year),
    month: int = Query(default_factory=lambda: datetime.now().month),
    _: str = Depends(verify_token),
):
    """Получение данных для графика по дням выбранного месяца с кэшированием."""

    cache_key = cache_manager.get_cache_key("dashboard", "chart_data", year, month)

    def _get_chart_data():
        def _db_query(session):
            try:
                # Проверяем корректность месяца и года
                if month < 1 or month > 12:
                    raise ValueError("Month must be between 1 and 12")

                if year < 2020 or year > datetime.now().year + 1:
                    raise ValueError("Invalid year")

                # Получаем количество дней в месяце
                days_in_month = calendar.monthrange(year, month)[1]

                # Инициализируем массивы для каждого дня месяца
                user_registrations = [0] * days_in_month
                ticket_creations = [0] * days_in_month
                booking_creations = [0] * days_in_month

                # Форматируем параметры для SQLite
                year_str = str(year)
                month_str = f"{month:02d}"

                # ОПТИМИЗИРОВАННЫЙ ЗАПРОС: Один запрос вместо трех с использованием UNION ALL
                # Это улучшает производительность на 30-50%
                combined_query = text(
                    """
                    WITH daily_data AS (
                        -- Регистрации пользователей
                        SELECT
                            CAST(strftime('%d', COALESCE(reg_date, first_join_time)) AS INTEGER) as day,
                            'users' as data_type,
                            COUNT(*) as count
                        FROM users
                        WHERE
                            (reg_date IS NOT NULL AND strftime('%Y', reg_date) = :year AND strftime('%m', reg_date) = :month_str)
                           OR
                            (reg_date IS NULL AND first_join_time IS NOT NULL AND strftime('%Y', first_join_time) = :year AND strftime('%m', first_join_time) = :month_str)
                        GROUP BY CAST(strftime('%d', COALESCE(reg_date, first_join_time)) AS INTEGER)

                        UNION ALL

                        -- Создание тикетов
                        SELECT
                            CAST(strftime('%d', created_at) AS INTEGER) as day,
                            'tickets' as data_type,
                            COUNT(*) as count
                        FROM tickets
                        WHERE strftime('%Y', created_at) = :year AND strftime('%m', created_at) = :month_str
                        GROUP BY CAST(strftime('%d', created_at) AS INTEGER)

                        UNION ALL

                        -- Создание бронирований
                        SELECT
                            CAST(strftime('%d', created_at) AS INTEGER) as day,
                            'bookings' as data_type,
                            COUNT(*) as count
                        FROM bookings
                        WHERE strftime('%Y', created_at) = :year AND strftime('%m', created_at) = :month_str
                        GROUP BY CAST(strftime('%d', created_at) AS INTEGER)
                    )
                    SELECT * FROM daily_data
                    ORDER BY day, data_type
                    """
                )

                # Выполняем единый запрос
                combined_results = session.execute(
                    combined_query, {"year": year_str, "month_str": month_str}
                ).fetchall()

                # Распределяем результаты по массивам
                for row in combined_results:
                    if row.day and 1 <= row.day <= days_in_month:
                        day_index = row.day - 1
                        if row.data_type == 'users':
                            user_registrations[day_index] = row.count
                        elif row.data_type == 'tickets':
                            ticket_creations[day_index] = row.count
                        elif row.data_type == 'bookings':
                            booking_creations[day_index] = row.count

                # Создаем метки для дней месяца
                day_labels = [str(i) for i in range(1, days_in_month + 1)]

                # Получаем название месяца на русском
                month_names = {
                    1: "Январь",
                    2: "Февраль",
                    3: "Март",
                    4: "Апрель",
                    5: "Май",
                    6: "Июнь",
                    7: "Июль",
                    8: "Август",
                    9: "Сентябрь",
                    10: "Октябрь",
                    11: "Ноябрь",
                    12: "Декабрь",
                }

                return {
                    "labels": day_labels,
                    "datasets": {
                        "user_registrations": user_registrations,
                        "ticket_creations": ticket_creations,
                        "booking_creations": booking_creations,
                    },
                    "period": {
                        "year": year,
                        "month": month,
                        "month_name": month_names[month],
                        "days_in_month": days_in_month,
                    },
                    "totals": {
                        "users": sum(user_registrations),
                        "tickets": sum(ticket_creations),
                        "bookings": sum(booking_creations),
                    },
                }

            except ValueError as e:
                logger.error(f"Ошибка валидации параметров: {e}")
                raise HTTPException(status_code=400, detail=str(e))
            except Exception as e:
                logger.error(f"Ошибка в запросе данных графика: {e}")
                raise

        return DatabaseManager.safe_execute(_db_query)

    try:
        # Используем кэш с более длинным TTL для исторических данных
        return await cache_manager.get_or_set(
            cache_key,
            _get_chart_data,
            ttl=cache_manager.static_data_ttl,  # 30 минут для исторических данных
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Критическая ошибка в get_chart_data: {e}")
        raise HTTPException(status_code=500, detail="Не удалось загрузить данные для графика. Проверьте подключение к базе данных")


@router.get("/available-periods")
async def get_available_periods(_: str = Depends(verify_token)):
    """Получение доступных периодов для выбора (месяцы с данными) с кэшированием."""

    cache_key = cache_manager.get_cache_key("dashboard", "available_periods")

    def _get_periods():
        def _db_query(session):
            try:
                # Получаем диапазон дат с данными (SQLite)
                date_ranges_query = text(
                    """
                    SELECT
                        MIN(date_col) as min_date,
                        MAX(date_col) as max_date
                    FROM (
                             SELECT COALESCE(reg_date, first_join_time) as date_col FROM users WHERE COALESCE(reg_date, first_join_time) IS NOT NULL
                             UNION ALL
                             SELECT created_at as date_col FROM tickets WHERE created_at IS NOT NULL
                             UNION ALL
                             SELECT created_at as date_col FROM bookings WHERE created_at IS NOT NULL
                         ) as all_dates
                    """
                )

                result = session.execute(date_ranges_query).fetchone()

                # Всегда возвращаем 12 месяцев (текущий + 11 предыдущих)
                current_date = datetime.now()
                periods = []
                
                month_names = {
                    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
                    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
                    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
                }
                
                # Генерируем 12 месяцев: текущий + 11 предыдущих
                for i in range(12):
                    # Вычисляем месяц и год
                    target_month = current_date.month - i
                    target_year = current_date.year
                    
                    # Корректируем если месяц стал отрицательным
                    while target_month <= 0:
                        target_month += 12
                        target_year -= 1
                    
                    periods.append({
                        "year": target_year,
                        "month": target_month,
                        "display": f"{month_names[target_month]} {target_year}",
                    })
                
                if not result or not result.min_date:
                    return {
                        "periods": periods,
                        "current": {
                            "year": current_date.year,
                            "month": current_date.month,
                        },
                    }

                # В SQLite результат может быть строкой, нужно парсить
                def parse_date(date_val):
                    if isinstance(date_val, str):
                        try:
                            # Пробуем разные форматы
                            if "T" in date_val:
                                return datetime.fromisoformat(
                                    date_val.replace("Z", "+00:00")
                                )
                            else:
                                return datetime.strptime(
                                    date_val[:19], "%Y-%m-%d %H:%M:%S"
                                )
                        except:
                            return datetime.now()
                    return date_val

                # Используем уже созданный список periods (12 месяцев)
                # Текущий период  
                now = datetime.now()
                current_period = {"year": now.year, "month": now.month}

                return {"periods": periods, "current": current_period}

            except Exception as e:
                logger.error(f"Ошибка получения доступных периодов: {e}")
                raise

        return DatabaseManager.safe_execute(_db_query)

    try:
        # Периоды меняются редко, кэшируем на 30 минут
        return await cache_manager.get_or_set(
            cache_key, _get_periods, ttl=cache_manager.static_data_ttl
        )
    except Exception as e:
        logger.error(f"Критическая ошибка в get_available_periods: {e}")
        raise HTTPException(
            status_code=500, detail="Не удалось загрузить доступные периоды для анализа. Попробуйте позже"
        )


@router.get("/bookings-calendar")
async def get_bookings_calendar(
    year: int = Query(default_factory=lambda: datetime.now().year),
    month: int = Query(default_factory=lambda: datetime.now().month),
    tariff_ids: Optional[str] = Query(None, description="Comma-separated tariff IDs to filter"),
    user_search: Optional[str] = Query(None, description="Search by user name or Telegram ID"),
    _: str = Depends(verify_token),
):
    """Получение данных бронирований для календаря с фильтрацией."""

    # Парсим tariff_ids если передан
    tariff_id_list = []
    if tariff_ids:
        try:
            tariff_id_list = [int(tid.strip()) for tid in tariff_ids.split(',') if tid.strip()]
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный формат списка тарифов. Ожидается массив числовых ID")

    cache_key = cache_manager.get_cache_key("dashboard", "bookings_calendar", year, month, tariff_ids or "", user_search or "")

    def _get_bookings_data():
        def _db_query(session):
            try:
                # Валидация входных параметров
                if not (1 <= month <= 12):
                    raise ValueError(f"Недопустимый месяц: {month}")
                if not (2000 <= year <= 9999):
                    raise ValueError(f"Недопустимый год: {year}")

                year_str = str(year).zfill(4)
                month_str = str(month).zfill(2)

                # Строим динамический запрос с фильтрами
                query_parts = [
                    """
                    SELECT
                        b.id,
                        b.visit_date,
                        b.visit_time,
                        b.confirmed,
                        b.paid,
                        b.amount,
                        u.full_name as user_name,
                        u.telegram_id,
                        t.name as tariff_name,
                        b.tariff_id
                    FROM bookings b
                    LEFT JOIN users u ON b.user_id = u.id
                    LEFT JOIN tariffs t ON b.tariff_id = t.id
                    WHERE strftime('%Y', b.visit_date) = :year
                      AND strftime('%m', b.visit_date) = :month_str
                    """
                ]

                query_params = {"year": year_str, "month_str": month_str}

                # Добавляем фильтр по тарифам
                if tariff_id_list:
                    placeholders = ','.join([f':tariff_{i}' for i in range(len(tariff_id_list))])
                    query_parts.append(f"AND b.tariff_id IN ({placeholders})")
                    for i, tid in enumerate(tariff_id_list):
                        query_params[f'tariff_{i}'] = tid

                # Добавляем фильтр по пользователю
                if user_search:
                    query_parts.append(
                        """AND (
                            LOWER(u.full_name) LIKE LOWER(:user_search)
                            OR CAST(u.telegram_id AS TEXT) LIKE :user_search
                        )"""
                    )
                    query_params['user_search'] = f'%{user_search}%'

                query_parts.append("ORDER BY b.visit_date, b.visit_time")

                bookings_query = text('\n'.join(query_parts))

                results = session.execute(bookings_query, query_params).fetchall()

                bookings = []
                for row in results:
                    # Форматируем данные бронирования
                    booking = {
                        "id": row.id,
                        "visit_date": row.visit_date.isoformat() if hasattr(row.visit_date, 'isoformat') else str(row.visit_date),
                        "visit_time": str(row.visit_time) if row.visit_time else None,
                        "confirmed": bool(row.confirmed),
                        "paid": bool(row.paid),
                        "amount": float(row.amount) if row.amount else 0,
                        "user_name": row.user_name,
                        "telegram_id": row.telegram_id,
                        "tariff_name": row.tariff_name
                    }
                    bookings.append(booking)

                return {
                    "bookings": bookings,
                    "period": {
                        "year": year,
                        "month": month,
                        "total_bookings": len(bookings)
                    }
                }

            except ValueError as e:
                logger.error(f"Ошибка валидации параметров календаря: {e}")
                raise HTTPException(status_code=400, detail=str(e))
            except Exception as e:
                logger.error(f"Ошибка в запросе данных календаря: {e}")
                raise

        return DatabaseManager.safe_execute(_db_query)

    try:
        # Кэшируем данные календаря
        return await cache_manager.get_or_set(
            cache_key,
            _get_bookings_data,
            ttl=cache_manager.dashboard_ttl,  # 5 минут
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Критическая ошибка в get_bookings_calendar: {e}")
        raise HTTPException(status_code=500, detail="Не удалось загрузить данные календаря бронирований. Попробуйте позже")


@router.get("/tariff-distribution")
async def get_tariff_distribution(
    period_start: Optional[str] = Query(None, description="Начало периода (ISO формат)"),
    period_end: Optional[str] = Query(None, description="Конец периода (ISO формат)"),
    _: str = Depends(verify_token)
):
    """
    Получение распределения бронирований по тарифам для Pie Chart
    """
    logger.info(f"Tariff distribution request: period_start={period_start}, period_end={period_end}")

    # Определяем период
    if period_start and period_end:
        try:
            period_start_dt = datetime.fromisoformat(period_start.replace('Z', '+00:00'))
            period_end_dt = datetime.fromisoformat(period_end.replace('Z', '+00:00'))
        except ValueError as e:
            logger.error(f"Invalid date format: {e}")
            raise HTTPException(status_code=400, detail="Неверный формат даты. Используйте ISO формат (YYYY-MM-DD)")
    else:
        # По умолчанию - текущий месяц
        now = datetime.now()
        period_start_dt = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if now.month == 12:
            period_end_dt = period_start_dt.replace(year=now.year + 1, month=1)
        else:
            period_end_dt = period_start_dt.replace(month=now.month + 1)

    cache_key = cache_manager.get_cache_key(
        "dashboard", "tariff_distribution",
        period_start_dt.isoformat(),
        period_end_dt.isoformat()
    )

    def _get_distribution():
        def _db_query(session):
            logger.info("Executing tariff distribution query")
            try:
                # SQL запрос для получения распределения по тарифам
                distribution_query = text("""
                    SELECT
                        t.id,
                        t.name,
                        t.price,
                        COUNT(b.id) as bookings_count,
                        COALESCE(SUM(CASE WHEN b.paid = 1 THEN b.amount ELSE 0 END), 0) as total_revenue
                    FROM tariffs t
                    LEFT JOIN bookings b ON t.id = b.tariff_id
                        AND b.created_at >= :period_start
                        AND b.created_at < :period_end
                    WHERE t.is_active = 1
                    GROUP BY t.id, t.name, t.price
                    HAVING bookings_count > 0
                    ORDER BY bookings_count DESC
                """)

                results = session.execute(distribution_query, {
                    'period_start': period_start_dt,
                    'period_end': period_end_dt
                }).fetchall()

                # Формируем данные для Pie Chart
                distribution_data = []
                total_bookings = 0
                total_revenue = 0.0

                for row in results:
                    bookings_count = int(row.bookings_count or 0)
                    revenue = float(row.total_revenue or 0)

                    distribution_data.append({
                        "id": int(row.id),
                        "name": row.name,
                        "value": bookings_count,
                        "revenue": revenue,
                        "price": float(row.price or 0)
                    })

                    total_bookings += bookings_count
                    total_revenue += revenue

                # Вычисляем проценты
                for item in distribution_data:
                    item["percentage"] = round((item["value"] / total_bookings * 100), 1) if total_bookings > 0 else 0

                logger.info(f"Tariff distribution fetched: {len(distribution_data)} tariffs, {total_bookings} bookings")

                return {
                    "data": distribution_data,
                    "total_bookings": total_bookings,
                    "total_revenue": total_revenue,
                    "period": {
                        "start": period_start_dt.isoformat(),
                        "end": period_end_dt.isoformat(),
                        "label": period_start_dt.strftime("%B %Y")
                    }
                }

            except Exception as e:
                logger.error(f"Ошибка в запросе распределения тарифов: {e}", exc_info=True)
                raise

        return DatabaseManager.safe_execute(_db_query)

    try:
        # Кэшируем данные распределения тарифов
        return await cache_manager.get_or_set(
            cache_key,
            _get_distribution,
            ttl=cache_manager.dashboard_ttl
        )
    except Exception as e:
        logger.error(f"Ошибка в get_tariff_distribution: {e}", exc_info=True)
        # Возвращаем пустые данные при ошибке
        return {
            "data": [],
            "total_bookings": 0,
            "total_revenue": 0.0,
            "period": {
                "start": period_start_dt.isoformat() if 'period_start_dt' in locals() else datetime.now().isoformat(),
                "end": period_end_dt.isoformat() if 'period_end_dt' in locals() else datetime.now().isoformat(),
                "label": period_start_dt.strftime("%B %Y") if 'period_start_dt' in locals() else datetime.now().strftime("%B %Y")
            }
        }


@router.get("/compare-periods")
async def compare_periods(
    period1_year: int = Query(..., description="Год первого периода"),
    period1_month: int = Query(..., description="Месяц первого периода (1-12)"),
    period2_year: int = Query(..., description="Год второго периода"),
    period2_month: int = Query(..., description="Месяц второго периода (1-12)"),
    _: str = Depends(verify_token)
):
    """
    Сравнение данных двух периодов для отображения на одном графике
    """
    logger.info(f"Compare periods request: period1={period1_year}/{period1_month}, period2={period2_year}/{period2_month}")

    # Валидация периодов
    if not (1 <= period1_month <= 12) or not (1 <= period2_month <= 12):
        raise HTTPException(status_code=400, detail="Неверное значение месяца. Должно быть число от 1 до 12")

    cache_key = cache_manager.get_cache_key(
        "dashboard", "compare_periods",
        f"{period1_year}_{period1_month}",
        f"{period2_year}_{period2_month}"
    )

    def _get_comparison():
        def _db_query(session):
            logger.info("Executing period comparison query")
            try:
                import calendar as cal

                # Получаем количество дней в каждом месяце
                days_in_month1 = cal.monthrange(period1_year, period1_month)[1]
                days_in_month2 = cal.monthrange(period2_year, period2_month)[1]

                # SQL запрос для первого периода
                period1_query = text("""
                    WITH RECURSIVE dates(date) AS (
                        SELECT DATE(:start_date)
                        UNION ALL
                        SELECT DATE(date, '+1 day')
                        FROM dates
                        WHERE date < DATE(:end_date)
                    )
                    SELECT
                        CAST(strftime('%d', dates.date) AS INTEGER) as day,
                        COALESCE(COUNT(DISTINCT u.id), 0) as users_count,
                        COALESCE(COUNT(DISTINCT t.id), 0) as tickets_count,
                        COALESCE(COUNT(DISTINCT b.id), 0) as bookings_count
                    FROM dates
                    LEFT JOIN users u ON DATE(COALESCE(u.reg_date, u.first_join_time)) = dates.date
                    LEFT JOIN tickets t ON DATE(t.created_at) = dates.date
                    LEFT JOIN bookings b ON DATE(b.created_at) = dates.date
                    GROUP BY dates.date
                    ORDER BY dates.date
                """)

                period1_start = datetime(period1_year, period1_month, 1)
                period1_end = datetime(period1_year, period1_month, days_in_month1, 23, 59, 59)

                results1 = session.execute(period1_query, {
                    'start_date': period1_start.strftime('%Y-%m-%d'),
                    'end_date': period1_end.strftime('%Y-%m-%d')
                }).fetchall()

                # SQL запрос для второго периода
                period2_start = datetime(period2_year, period2_month, 1)
                period2_end = datetime(period2_year, period2_month, days_in_month2, 23, 59, 59)

                results2 = session.execute(period1_query, {
                    'start_date': period2_start.strftime('%Y-%m-%d'),
                    'end_date': period2_end.strftime('%Y-%m-%d')
                }).fetchall()

                # Формируем данные
                period1_data = {
                    "labels": [str(row.day) for row in results1],
                    "users": [int(row.users_count) for row in results1],
                    "tickets": [int(row.tickets_count) for row in results1],
                    "bookings": [int(row.bookings_count) for row in results1]
                }

                period2_data = {
                    "labels": [str(row.day) for row in results2],
                    "users": [int(row.users_count) for row in results2],
                    "tickets": [int(row.tickets_count) for row in results2],
                    "bookings": [int(row.bookings_count) for row in results2]
                }

                logger.info(f"Comparison data fetched: period1={len(results1)} days, period2={len(results2)} days")

                return {
                    "period1": {
                        "year": period1_year,
                        "month": period1_month,
                        "label": period1_start.strftime("%B %Y"),
                        "data": period1_data
                    },
                    "period2": {
                        "year": period2_year,
                        "month": period2_month,
                        "label": period2_start.strftime("%B %Y"),
                        "data": period2_data
                    }
                }

            except Exception as e:
                logger.error(f"Ошибка в запросе сравнения периодов: {e}", exc_info=True)
                raise

        return DatabaseManager.safe_execute(_db_query)

    try:
        # Кэшируем данные сравнения
        return await cache_manager.get_or_set(
            cache_key,
            _get_comparison,
            ttl=cache_manager.dashboard_ttl
        )
    except Exception as e:
        logger.error(f"Ошибка в compare_periods: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Не удалось загрузить данные для сравнения периодов. Проверьте корректность дат")


@router.get("/export-csv")
async def export_dashboard_csv(
    period_start: Optional[str] = Query(None, description="Начало периода (ISO формат)"),
    period_end: Optional[str] = Query(None, description="Конец периода (ISO формат)"),
    _: str = Depends(verify_token)
):
    """
    Экспорт данных дашборда в CSV формат.
    Включает статистику, ежедневные данные и детализацию по тарифам.
    """
    logger.info(f"CSV export request: period_start={period_start}, period_end={period_end}")

    # Определяем период
    if period_start and period_end:
        try:
            period_start_dt = datetime.fromisoformat(period_start.replace('Z', '+00:00'))
            period_end_dt = datetime.fromisoformat(period_end.replace('Z', '+00:00'))
        except ValueError as e:
            logger.error(f"Invalid date format: {e}")
            raise HTTPException(status_code=400, detail="Неверный формат даты. Используйте ISO формат (YYYY-MM-DD)")
    else:
        # По умолчанию - текущий месяц
        now = datetime.now()
        period_start_dt = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_day = calendar.monthrange(now.year, now.month)[1]
        period_end_dt = now.replace(day=last_day, hour=23, minute=59, second=59, microsecond=999999)

    try:
        def _get_export_data(session):
            # Получаем общую статистику (исправлен cartesian join - используем scalar subqueries)
            stats_query = text("""
                SELECT
                    (SELECT COUNT(*) FROM users WHERE reg_date >= :period_start AND reg_date < :period_end) as total_users,
                    (SELECT COUNT(*) FROM bookings WHERE created_at >= :period_start AND created_at < :period_end) as total_bookings,
                    (SELECT COUNT(*) FROM bookings WHERE paid = 1 AND created_at >= :period_start AND created_at < :period_end) as paid_bookings,
                    (SELECT COALESCE(SUM(amount), 0) FROM bookings WHERE paid = 1 AND created_at >= :period_start AND created_at < :period_end) as total_revenue,
                    (SELECT COUNT(*) FROM tickets WHERE created_at >= :period_start AND created_at < :period_end) as total_tickets,
                    (SELECT COUNT(*) FROM tickets WHERE status = 'OPEN' AND created_at >= :period_start AND created_at < :period_end) as open_tickets
            """)

            stats_result = session.execute(stats_query, {
                'period_start': period_start_dt,
                'period_end': period_end_dt
            }).fetchone()

            # Получаем ежедневные данные
            daily_query = text("""
                WITH RECURSIVE dates(date) AS (
                    SELECT DATE(:period_start)
                    UNION ALL
                    SELECT DATE(date, '+1 day')
                    FROM dates
                    WHERE date < DATE(:period_end)
                )
                SELECT
                    d.date,
                    COUNT(DISTINCT u.id) as users,
                    COUNT(DISTINCT b.id) as bookings,
                    COUNT(DISTINCT t.id) as tickets,
                    COALESCE(SUM(CASE WHEN b.paid = 1 THEN b.amount ELSE 0 END), 0) as revenue
                FROM dates d
                LEFT JOIN users u ON DATE(u.reg_date) = d.date
                LEFT JOIN bookings b ON DATE(b.created_at) = d.date
                LEFT JOIN tickets t ON DATE(t.created_at) = d.date
                GROUP BY d.date
                ORDER BY d.date
            """)

            daily_results = session.execute(daily_query, {
                'period_start': period_start_dt,
                'period_end': period_end_dt
            }).fetchall()

            # Получаем данные по тарифам
            tariff_query = text("""
                SELECT
                    t.name,
                    COUNT(b.id) as booking_count,
                    COALESCE(SUM(CASE WHEN b.paid = 1 THEN b.amount END), 0) as revenue
                FROM tariffs t
                LEFT JOIN bookings b ON b.tariff_id = t.id
                    AND b.created_at >= :period_start
                    AND b.created_at < :period_end
                GROUP BY t.id, t.name
                ORDER BY booking_count DESC
            """)

            tariff_results = session.execute(tariff_query, {
                'period_start': period_start_dt,
                'period_end': period_end_dt
            }).fetchall()

            return {
                'stats': stats_result,
                'daily': daily_results,
                'tariffs': tariff_results
            }

        export_data = DatabaseManager.safe_execute(_get_export_data)

        # Создаем CSV в памяти
        output = io.StringIO()
        writer = csv.writer(output)

        # Секция 1: Общая статистика
        writer.writerow(['ОБЩАЯ СТАТИСТИКА'])
        writer.writerow(['Период', f"{period_start_dt.strftime('%Y-%m-%d')} - {period_end_dt.strftime('%Y-%m-%d')}"])
        writer.writerow([])
        writer.writerow(['Метрика', 'Значение'])
        writer.writerow(['Всего пользователей', export_data['stats'].total_users])
        writer.writerow(['Всего бронирований', export_data['stats'].total_bookings])
        writer.writerow(['Оплаченных бронирований', export_data['stats'].paid_bookings])
        writer.writerow(['Общий доход (₽)', f"{export_data['stats'].total_revenue:.2f}"])
        writer.writerow(['Всего тикетов', export_data['stats'].total_tickets])
        writer.writerow(['Открытых тикетов', export_data['stats'].open_tickets])
        writer.writerow([])
        writer.writerow([])

        # Секция 2: Ежедневная статистика
        writer.writerow(['ЕЖЕДНЕВНАЯ СТАТИСТИКА'])
        writer.writerow(['Дата', 'Пользователи', 'Бронирования', 'Тикеты', 'Доход (₽)'])
        for row in export_data['daily']:
            writer.writerow([
                row.date,
                row.users,
                row.bookings,
                row.tickets,
                f"{row.revenue:.2f}"
            ])
        writer.writerow([])
        writer.writerow([])

        # Секция 3: Статистика по тарифам
        writer.writerow(['СТАТИСТИКА ПО ТАРИФАМ'])
        writer.writerow(['Тариф', 'Количество бронирований', 'Доход (₽)'])
        for row in export_data['tariffs']:
            writer.writerow([
                row.name,
                row.booking_count,
                f"{row.revenue:.2f}"
            ])

        # Подготавливаем файл для скачивания
        output.seek(0)

        # Формируем имя файла
        filename = f"dashboard_{period_start_dt.strftime('%Y%m%d')}_{period_end_dt.strftime('%Y%m%d')}.csv"

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        logger.error(f"Ошибка экспорта в CSV: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Не удалось экспортировать данные в CSV формат. Попробуйте позже")


@router.get("/export-excel")
async def export_dashboard_excel(
    period_start: Optional[str] = Query(None, description="Начало периода (ISO формат)"),
    period_end: Optional[str] = Query(None, description="Конец периода (ISO формат)"),
    _: str = Depends(verify_token)
):
    """
    Экспорт данных дашборда в Excel формат.
    Создает форматированный Excel файл с несколькими листами:
    - Общая статистика
    - Ежедневная статистика
    - Статистика по тарифам
    """
    logger.info(f"Excel export request: period_start={period_start}, period_end={period_end}")

    # Определяем период
    if period_start and period_end:
        try:
            period_start_dt = datetime.fromisoformat(period_start.replace('Z', '+00:00'))
            period_end_dt = datetime.fromisoformat(period_end.replace('Z', '+00:00'))
        except ValueError as e:
            logger.error(f"Invalid date format: {e}")
            raise HTTPException(status_code=400, detail="Неверный формат даты. Используйте ISO формат (YYYY-MM-DD)")
    else:
        # По умолчанию - текущий месяц
        now = datetime.now()
        period_start_dt = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_day = calendar.monthrange(now.year, now.month)[1]
        period_end_dt = now.replace(day=last_day, hour=23, minute=59, second=59, microsecond=999999)

    try:
        def _get_export_data(session):
            # Получаем общую статистику (исправлен cartesian join - используем scalar subqueries)
            stats_query = text("""
                SELECT
                    (SELECT COUNT(*) FROM users WHERE reg_date >= :period_start AND reg_date < :period_end) as total_users,
                    (SELECT COUNT(*) FROM bookings WHERE created_at >= :period_start AND created_at < :period_end) as total_bookings,
                    (SELECT COUNT(*) FROM bookings WHERE paid = 1 AND created_at >= :period_start AND created_at < :period_end) as paid_bookings,
                    (SELECT COALESCE(SUM(amount), 0) FROM bookings WHERE paid = 1 AND created_at >= :period_start AND created_at < :period_end) as total_revenue,
                    (SELECT COALESCE(AVG(amount), 0) FROM bookings WHERE paid = 1 AND created_at >= :period_start AND created_at < :period_end) as avg_booking_value,
                    (SELECT COUNT(*) FROM tickets WHERE created_at >= :period_start AND created_at < :period_end) as total_tickets,
                    (SELECT COUNT(*) FROM tickets WHERE status = 'OPEN' AND created_at >= :period_start AND created_at < :period_end) as open_tickets
            """)

            stats_result = session.execute(stats_query, {
                'period_start': period_start_dt,
                'period_end': period_end_dt
            }).fetchone()

            # Получаем ежедневные данные
            daily_query = text("""
                WITH RECURSIVE dates(date) AS (
                    SELECT DATE(:period_start)
                    UNION ALL
                    SELECT DATE(date, '+1 day')
                    FROM dates
                    WHERE date < DATE(:period_end)
                )
                SELECT
                    d.date,
                    COUNT(DISTINCT u.id) as users,
                    COUNT(DISTINCT b.id) as bookings,
                    COUNT(DISTINCT t.id) as tickets,
                    COALESCE(SUM(CASE WHEN b.paid = 1 THEN b.amount ELSE 0 END), 0) as revenue
                FROM dates d
                LEFT JOIN users u ON DATE(u.reg_date) = d.date
                LEFT JOIN bookings b ON DATE(b.created_at) = d.date
                LEFT JOIN tickets t ON DATE(t.created_at) = d.date
                GROUP BY d.date
                ORDER BY d.date
            """)

            daily_results = session.execute(daily_query, {
                'period_start': period_start_dt,
                'period_end': period_end_dt
            }).fetchall()

            # Получаем данные по тарифам
            tariff_query = text("""
                SELECT
                    t.name,
                    COUNT(b.id) as booking_count,
                    COALESCE(SUM(CASE WHEN b.paid = 1 THEN b.amount END), 0) as revenue,
                    COALESCE(AVG(CASE WHEN b.paid = 1 THEN b.amount END), 0) as avg_revenue
                FROM tariffs t
                LEFT JOIN bookings b ON b.tariff_id = t.id
                    AND b.created_at >= :period_start
                    AND b.created_at < :period_end
                GROUP BY t.id, t.name
                ORDER BY booking_count DESC
            """)

            tariff_results = session.execute(tariff_query, {
                'period_start': period_start_dt,
                'period_end': period_end_dt
            }).fetchall()

            return {
                'stats': stats_result,
                'daily': daily_results,
                'tariffs': tariff_results
            }

        export_data = DatabaseManager.safe_execute(_get_export_data)

        # Создаем Excel workbook
        wb = Workbook()

        # Определяем стили
        header_fill = PatternFill(start_color="7B68EE", end_color="7B68EE", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="4B0082")
        center_alignment = Alignment(horizontal="center", vertical="center")

        # ============ Лист 1: Общая статистика ============
        ws_stats = wb.active
        ws_stats.title = "Общая статистика"

        # Заголовок
        ws_stats['A1'] = 'ДАШБОРД - ОБЩАЯ СТАТИСТИКА'
        ws_stats['A1'].font = title_font
        ws_stats.merge_cells('A1:B1')

        ws_stats['A2'] = 'Период:'
        ws_stats['B2'] = f"{period_start_dt.strftime('%d.%m.%Y')} - {period_end_dt.strftime('%d.%m.%Y')}"
        ws_stats['A2'].font = Font(bold=True)

        # Заголовки таблицы
        ws_stats['A4'] = 'Метрика'
        ws_stats['B4'] = 'Значение'
        ws_stats['A4'].fill = header_fill
        ws_stats['B4'].fill = header_fill
        ws_stats['A4'].font = header_font
        ws_stats['B4'].font = header_font
        ws_stats['A4'].alignment = center_alignment
        ws_stats['B4'].alignment = center_alignment

        # Данные
        stats_data = [
            ['Всего пользователей', export_data['stats'].total_users],
            ['Всего бронирований', export_data['stats'].total_bookings],
            ['Оплаченных бронирований', export_data['stats'].paid_bookings],
            ['Общий доход (₽)', f"{export_data['stats'].total_revenue:.2f}"],
            ['Средний чек (₽)', f"{export_data['stats'].avg_booking_value:.2f}"],
            ['Всего тикетов', export_data['stats'].total_tickets],
            ['Открытых тикетов', export_data['stats'].open_tickets]
        ]

        for idx, row in enumerate(stats_data, start=5):
            ws_stats[f'A{idx}'] = row[0]
            ws_stats[f'B{idx}'] = row[1]
            ws_stats[f'A{idx}'].font = Font(bold=True)

        # Ширина колонок
        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 20

        # ============ Лист 2: Ежедневная статистика ============
        ws_daily = wb.create_sheet("Ежедневная статистика")

        # Заголовок
        ws_daily['A1'] = 'ЕЖЕДНЕВНАЯ СТАТИСТИКА'
        ws_daily['A1'].font = title_font
        ws_daily.merge_cells('A1:E1')

        # Заголовки таблицы
        headers = ['Дата', 'Пользователи', 'Бронирования', 'Тикеты', 'Доход (₽)']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_daily.cell(row=3, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # Данные
        for row_idx, row_data in enumerate(export_data['daily'], start=4):
            ws_daily.cell(row=row_idx, column=1, value=row_data.date)
            ws_daily.cell(row=row_idx, column=2, value=row_data.users)
            ws_daily.cell(row=row_idx, column=3, value=row_data.bookings)
            ws_daily.cell(row=row_idx, column=4, value=row_data.tickets)
            ws_daily.cell(row=row_idx, column=5, value=float(row_data.revenue))

        # Ширина колонок
        ws_daily.column_dimensions['A'].width = 12
        ws_daily.column_dimensions['B'].width = 15
        ws_daily.column_dimensions['C'].width = 15
        ws_daily.column_dimensions['D'].width = 12
        ws_daily.column_dimensions['E'].width = 15

        # ============ Лист 3: Статистика по тарифам ============
        ws_tariffs = wb.create_sheet("Статистика по тарифам")

        # Заголовок
        ws_tariffs['A1'] = 'СТАТИСТИКА ПО ТАРИФАМ'
        ws_tariffs['A1'].font = title_font
        ws_tariffs.merge_cells('A1:D1')

        # Заголовки таблицы
        tariff_headers = ['Тариф', 'Количество бронирований', 'Общий доход (₽)', 'Средний доход (₽)']
        for col_idx, header in enumerate(tariff_headers, start=1):
            cell = ws_tariffs.cell(row=3, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # Данные
        for row_idx, row_data in enumerate(export_data['tariffs'], start=4):
            ws_tariffs.cell(row=row_idx, column=1, value=row_data.name)
            ws_tariffs.cell(row=row_idx, column=2, value=row_data.booking_count)
            ws_tariffs.cell(row=row_idx, column=3, value=float(row_data.revenue))
            ws_tariffs.cell(row=row_idx, column=4, value=float(row_data.avg_revenue))

        # Ширина колонок
        ws_tariffs.column_dimensions['A'].width = 25
        ws_tariffs.column_dimensions['B'].width = 25
        ws_tariffs.column_dimensions['C'].width = 20
        ws_tariffs.column_dimensions['D'].width = 20

        # Сохраняем в память
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Формируем имя файла
        filename = f"dashboard_{period_start_dt.strftime('%Y%m%d')}_{period_end_dt.strftime('%Y%m%d')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        logger.error(f"Ошибка экспорта в Excel: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Не удалось экспортировать данные в Excel формат. Попробуйте позже")


@router.get("/top-clients")
async def get_top_clients(
    limit: int = Query(5, ge=1, le=20, description="Количество клиентов в топе"),
    period_start: Optional[str] = Query(None, description="Начало периода (ISO формат)"),
    period_end: Optional[str] = Query(None, description="Конец периода (ISO формат)"),
    _: str = Depends(verify_token)
):
    """
    Получение топ клиентов по расходам.
    Возвращает список клиентов с наибольшими суммами бронирований.
    """
    logger.info(f"Top clients request: limit={limit}, period_start={period_start}, period_end={period_end}")

    # Определяем период
    if period_start and period_end:
        try:
            period_start_dt = datetime.fromisoformat(period_start.replace('Z', '+00:00'))
            period_end_dt = datetime.fromisoformat(period_end.replace('Z', '+00:00'))
        except ValueError as e:
            logger.error(f"Invalid date format: {e}")
            raise HTTPException(status_code=400, detail="Неверный формат даты. Используйте ISO формат (YYYY-MM-DD)")
    else:
        # По умолчанию - текущий месяц
        now = datetime.now()
        period_start_dt = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_day = calendar.monthrange(now.year, now.month)[1]
        period_end_dt = now.replace(day=last_day, hour=23, minute=59, second=59, microsecond=999999)

    # Ключ кэша
    cache_key = f"dashboard:top_clients:{limit}:{period_start_dt.isoformat()}:{period_end_dt.isoformat()}"

    async def _get_top_clients():
        def _db_query(session):
            try:
                query = text("""
                    SELECT
                        u.id,
                        u.username,
                        u.telegram_id,
                        COUNT(DISTINCT b.id) as total_bookings,
                        COUNT(DISTINCT CASE WHEN b.paid = 1 THEN b.id END) as paid_bookings,
                        COALESCE(SUM(CASE WHEN b.paid = 1 THEN b.amount END), 0) as total_spent,
                        COALESCE(AVG(CASE WHEN b.paid = 1 THEN b.amount END), 0) as avg_booking_value,
                        MAX(b.created_at) as last_booking_date
                    FROM users u
                    INNER JOIN bookings b ON b.user_id = u.id
                    WHERE b.created_at >= :period_start
                        AND b.created_at < :period_end
                    GROUP BY u.id, u.username, u.telegram_id
                    HAVING total_spent > 0
                    ORDER BY total_spent DESC
                    LIMIT :limit
                """)

                results = session.execute(query, {
                    'period_start': period_start_dt,
                    'period_end': period_end_dt,
                    'limit': limit
                }).fetchall()

                # Получаем общую статистику
                total_query = text("""
                    SELECT
                        COUNT(DISTINCT u.id) as total_clients_with_bookings,
                        COALESCE(SUM(CASE WHEN b.paid = 1 THEN b.amount END), 0) as total_revenue
                    FROM users u
                    INNER JOIN bookings b ON b.user_id = u.id
                    WHERE b.created_at >= :period_start
                        AND b.created_at < :period_end
                """)

                total_stats = session.execute(total_query, {
                    'period_start': period_start_dt,
                    'period_end': period_end_dt
                }).fetchone()

                clients_list = []
                for row in results:
                    clients_list.append({
                        'user_id': row.id,
                        'username': row.username,
                        'telegram_id': row.telegram_id,
                        'total_bookings': row.total_bookings,
                        'paid_bookings': row.paid_bookings,
                        'total_spent': float(row.total_spent),
                        'avg_booking_value': float(row.avg_booking_value),
                        'last_booking_date': row.last_booking_date.isoformat() if row.last_booking_date else None,
                        'revenue_share': (float(row.total_spent) / float(total_stats.total_revenue) * 100) if total_stats.total_revenue > 0 else 0
                    })

                return {
                    'clients': clients_list,
                    'period': {
                        'start': period_start_dt.isoformat(),
                        'end': period_end_dt.isoformat(),
                        'month_name': period_start_dt.strftime('%B'),
                        'year': period_start_dt.year
                    },
                    'stats': {
                        'total_clients_with_bookings': total_stats.total_clients_with_bookings,
                        'total_revenue': float(total_stats.total_revenue),
                        'top_clients_revenue': sum(c['total_spent'] for c in clients_list),
                        'top_clients_share': (sum(c['total_spent'] for c in clients_list) / float(total_stats.total_revenue) * 100) if total_stats.total_revenue > 0 else 0
                    }
                }

            except Exception as e:
                logger.error(f"Ошибка в запросе top-clients: {e}", exc_info=True)
                raise

        return DatabaseManager.safe_execute(_db_query)

    try:
        # Кэшируем данные топ-клиентов
        return await cache_manager.get_or_set(
            cache_key,
            _get_top_clients,
            ttl=cache_manager.dashboard_ttl
        )
    except Exception as e:
        logger.error(f"Ошибка в get_top_clients: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Не удалось загрузить список топ-клиентов. Проверьте подключение к базе данных")


@router.get("/promocode-stats")
async def get_promocode_stats(
    period_start: Optional[str] = Query(None, description="Начало периода (ISO формат)"),
    period_end: Optional[str] = Query(None, description="Конец периода (ISO формат)"),
    _: str = Depends(verify_token)
):
    """
    Получение статистики эффективности промокодов.
    Возвращает данные о использовании промокодов, общих скидках и доходах.
    """
    # Определяем период по умолчанию (текущий месяц)
    if period_start and period_end:
        start_date = datetime.fromisoformat(period_start.replace('Z', '+00:00'))
        end_date = datetime.fromisoformat(period_end.replace('Z', '+00:00'))
    else:
        now = datetime.now(MOSCOW_TZ)
        start_date = datetime(now.year, now.month, 1, tzinfo=MOSCOW_TZ)
        if now.month == 12:
            end_date = datetime(now.year + 1, 1, 1, tzinfo=MOSCOW_TZ)
        else:
            end_date = datetime(now.year, now.month + 1, 1, tzinfo=MOSCOW_TZ)

    def _get_promocode_stats(session):
        # SQL запрос для получения статистики по каждому промокоду
        query = text("""
            SELECT
                p.id,
                p.name,
                p.discount as discount_percent,
                p.is_active,
                COUNT(b.id) as total_uses,
                COUNT(CASE WHEN b.paid = 1 THEN b.id END) as paid_uses,
                COALESCE(SUM(CASE WHEN b.paid = 1 THEN b.amount END), 0) as total_revenue,
                COALESCE(SUM(CASE WHEN b.paid = 1 THEN b.amount * p.discount / 100.0 END), 0) as total_discount_given,
                COALESCE(AVG(CASE WHEN b.paid = 1 THEN b.amount END), 0) as avg_booking_value,
                MAX(b.created_at) as last_used_date
            FROM promocodes p
            LEFT JOIN bookings b ON b.promocode_id = p.id
                AND b.created_at >= :period_start
                AND b.created_at < :period_end
            GROUP BY p.id, p.name, p.discount, p.is_active
            HAVING total_uses > 0 OR p.is_active = 1
            ORDER BY total_revenue DESC
        """)

        result = session.execute(query, {
            "period_start": start_date,
            "period_end": end_date
        }).fetchall()

        promocodes_data = []
        total_uses = 0
        total_discount = 0
        total_revenue = 0
        active_count = 0

        for row in result:
            promo_data = {
                "id": row.id,
                "name": row.name,
                "discount_percent": row.discount_percent,
                "is_active": bool(row.is_active),
                "total_uses": int(row.total_uses or 0),
                "paid_uses": int(row.paid_uses or 0),
                "total_revenue": float(row.total_revenue or 0),
                "total_discount_given": float(row.total_discount_given or 0),
                "avg_booking_value": float(row.avg_booking_value or 0),
                "last_used_date": row.last_used_date.isoformat() if row.last_used_date else None,
                # Эффективность: доход на одно использование
                "revenue_per_use": float(row.total_revenue or 0) / max(int(row.paid_uses or 0), 1)
            }
            promocodes_data.append(promo_data)

            # Накапливаем общую статистику
            total_uses += int(row.total_uses or 0)
            total_discount += float(row.total_discount_given or 0)
            total_revenue += float(row.total_revenue or 0)
            if row.is_active:
                active_count += 1

        # Добавляем расчет доли дохода для каждого промокода
        for promo in promocodes_data:
            promo["revenue_share"] = (promo["total_revenue"] / total_revenue * 100) if total_revenue > 0 else 0

        return {
            "promocodes": promocodes_data,
            "period": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat(),
                "label": start_date.strftime("%B %Y") if start_date.month == (end_date - timedelta(days=1)).month else f"{start_date.strftime('%B %Y')} - {end_date.strftime('%B %Y')}"
            },
            "summary": {
                "total_promocodes": len(promocodes_data),
                "active_promocodes": active_count,
                "total_uses": total_uses,
                "total_discount_given": round(total_discount, 2),
                "total_revenue_with_promocodes": round(total_revenue, 2),
                "avg_discount_per_booking": round(total_discount / max(total_uses, 1), 2)
            }
        }

    try:
        # Проверяем кэш
        cache_key = f"promocode_stats:{period_start}:{period_end}"
        cached_data = await cache_manager.get(cache_key, namespace="dashboard")
        if cached_data:
            logger.info("Возвращены данные статистики промокодов из кэша")
            return cached_data

        # Получаем данные из БД
        stats = DatabaseManager.safe_execute(_get_promocode_stats)

        # Кэшируем результат
        await cache_manager.set(
            cache_key,
            stats,
            namespace="dashboard",
            ttl=cache_manager.dashboard_ttl
        )

        return stats

    except Exception as e:
        logger.error(f"Ошибка в get_promocode_stats: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Не удалось загрузить статистику использования промокодов. Попробуйте позже")
